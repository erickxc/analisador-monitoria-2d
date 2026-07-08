"""
Interface principal do Monitor (2D Consultores | Monitores).

Integra o motor de análise (analise_funil.py), o construtor de relatórios
personalizados (pivot_builder.py) e os gráficos de dispersão (graficos.py)
em uma única janela Tkinter com abas:

  Configurações -> Relatório Padrão -> Relatórios Personalizados -> Gráficos -> Perfil
"""

import os
import re
import sys
import json
import queue
import logging
import threading
import traceback
import webbrowser
from datetime import datetime

import tkinter as tk
import tkinter.font as tkfont
from tkinter import ttk, filedialog, messagebox

import recursos
from recursos import (
    CAMINHO_LOGO, CAMINHO_LOGO_ICO, CAMINHO_LOGO_ICONE, NOME_SISTEMA, NOME_EMPRESA, TITULO_JANELA,
    VERSAO_ATUAL, pasta_base_execucao,
)
import perfil
import atualizacoes

try:
    from tkinterdnd2 import TkinterDnD
    JANELA_BASE = TkinterDnD.Tk
except ImportError:
    JANELA_BASE = tk.Tk

# As dependências pesadas (pandas, openpyxl, matplotlib e os módulos do motor
# de análise) são importadas em _importar_dependencias_pesadas(), chamada
# pela tela de splash — assim a janela inicial aparece imediatamente, sem
# esperar essas bibliotecas carregarem. Os nomes abaixo são preenchidos nesse
# momento e usados pelo restante do arquivo (todos referenciados dentro de
# métodos/funções, nunca no corpo de uma classe, então a resolução do nome
# só acontece quando o código já rodou).
pd = None
Workbook = None
Font = PatternFill = Alignment = None
get_column_letter = None
ImagemExcel = None
af = pb = gf = None
sv_ttk = None


def _passo_importar_pandas():
    global pd
    import pandas as _pd
    pd = _pd


def _passo_importar_openpyxl():
    global Workbook, Font, PatternFill, Alignment, get_column_letter, ImagemExcel
    from openpyxl import Workbook as _Workbook
    from openpyxl.styles import Font as _Font, PatternFill as _PatternFill, Alignment as _Alignment
    from openpyxl.utils import get_column_letter as _get_column_letter
    from openpyxl.drawing.image import Image as _ImagemExcel
    Workbook = _Workbook
    Font, PatternFill, Alignment = _Font, _PatternFill, _Alignment
    get_column_letter = _get_column_letter
    ImagemExcel = _ImagemExcel


def _passo_importar_motor():
    global af
    import analise_funil as _af
    af = _af


def _passo_importar_construtor_relatorios():
    global pb
    import pivot_builder as _pb
    pb = _pb


def _passo_importar_graficos():
    global gf
    import graficos as _gf
    gf = _gf


def _passo_importar_tema_visual():
    global sv_ttk
    import sv_ttk as _sv_ttk
    sv_ttk = _sv_ttk


def _passo_verificar_banco_local():
    perfil.carregar_perfil()  # exercita criação/leitura real do banco SQLite local


def _passo_verificar_arquivos_sistema():
    if not os.path.exists(CAMINHO_LOGO):
        raise FileNotFoundError(f"logo não encontrada em {CAMINHO_LOGO}")


def _passo_limpar_logs_antigos(dias=30):
    pasta_logs = os.path.join(pasta_base_execucao(), "logs")
    if not os.path.isdir(pasta_logs):
        return
    limite = datetime.now().timestamp() - dias * 86400
    for nome_arquivo in os.listdir(pasta_logs):
        caminho = os.path.join(pasta_logs, nome_arquivo)
        try:
            if os.path.isfile(caminho) and os.path.getmtime(caminho) < limite:
                os.remove(caminho)
        except OSError:
            pass


def construir_etapas_preparacao():
    """
    Sequência de verificações reais executadas pela splash antes de abrir a
    janela principal (cada etapa realmente faz o que o texto diz — não é só
    enfeite visual).
    """
    return [
        ("Verificando bibliotecas de dados (pandas)...", _passo_importar_pandas),
        ("Verificando geração de planilhas (openpyxl)...", _passo_importar_openpyxl),
        ("Carregando motor de análise...", _passo_importar_motor),
        ("Carregando construtor de relatórios personalizados...", _passo_importar_construtor_relatorios),
        ("Verificando geração de gráficos (matplotlib)...", _passo_importar_graficos),
        ("Aplicando tema visual...", _passo_importar_tema_visual),
        ("Verificando banco de dados local (perfil)...", _passo_verificar_banco_local),
        ("Verificando arquivos de sistema (logo, ícones)...", _passo_verificar_arquivos_sistema),
        ("Limpando arquivos de log antigos (+30 dias)...", _passo_limpar_logs_antigos),
    ]


COR_CABECALHO = "1F4E78"
COR_ACCENT = f"#{COR_CABECALHO}"  # mesma cor de destaque usada nos cabeçalhos do Excel

# Catálogo de relatórios prontos oferecidos na aba "Relatório Padrão", agrupados
# por categoria para facilitar a leitura. Cada item mapeia um título de negócio
# para as chaves internas já calculadas por analise_funil.gerar_analises_completas.
CATALOGO_RELATORIOS = [
    ("Vendas", [
        ("top_clientes", "Venda por Cliente (Top Clientes)"),
        ("top_fabricantes", "Venda por Fabricante (Top Fabricantes)"),
        ("top_produtos", "Venda por Produto (Top Produtos)"),
    ]),
    ("Segmentação e Poder de Compra", [
        ("abc", "Faturamento e Segmentação de Clientes (ABC)"),
        ("poder_compra_clientes", "Poder de Compra por Cliente (Renúncia)"),
        ("migracao_abc", "Migração de Clientes entre Faixas"),
    ]),
    ("Tendências e Alertas", [
        ("evolucao_produtos", "Tendência de Produtos"),
        ("alertas_queda", "Alertas de Queda Consecutiva"),
        ("erosao_clientes", "Erosão de Clientes por Produto"),
    ]),
    ("Boletins", [
        ("produtos_em_alta", "Boletim: Produtos em Alta"),
        ("produtos_em_queda", "Boletim: Produtos em Queda"),
        ("clientes_queda_qtd", "Boletim: Clientes em Queda de Quantidade"),
        ("correlacao_produto_cliente", "Boletim: Correlação Produto x Cliente"),
        ("impacto_financeiro_churn", "Boletim: Impacto Financeiro do Churn"),
    ]),
]


def _preparar_logger():
    pasta_logs = recursos.pasta_logs()
    nome_arquivo = f"analise_funil_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    caminho_log = os.path.join(pasta_logs, nome_arquivo)

    logger = logging.getLogger("analise_funil_app")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    manipulador = logging.FileHandler(caminho_log, encoding="utf-8")
    manipulador.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logger.addHandler(manipulador)
    return logger, caminho_log


def _formatar_moeda_br(valor):
    return "R$ " + f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


class AplicacaoAnaliseFunil(JANELA_BASE):
    def __init__(self):
        super().__init__()
        self.title(TITULO_JANELA)
        self.geometry("1360x880")
        self._definir_icone_janela()
        sv_ttk.set_theme("light")
        self._aplicar_estilos_customizados()

        self.logger, self.caminho_log = _preparar_logger()
        self.logger.info("Aplicação iniciada.")

        self.perfil = perfil.carregar_perfil()

        # Estado compartilhado entre as abas
        self.df = None
        self.resultados_analise = None
        self.granularidade_referencia = "Mensal"
        self.relatorios_personalizados = {}
        self.estado_clientes = {}   # cliente -> True se EXCLUÍDO
        self.estado_produtos = {}   # produto -> True se CONSIDERADO
        self.thread_em_execucao = False
        self._thread_geracao = None

        self.fila_eventos = queue.Queue()

        self.tamanho_fonte_base = self.perfil["tamanho_fonte"]
        self.fator_zoom = 1.0

        self._montar_interface()
        self._aplicar_zoom(None)
        self._id_after_fila = self.after(150, self._bombear_fila_eventos)
        self.protocol("WM_DELETE_WINDOW", self._ao_fechar_janela)

        threading.Thread(target=self._verificar_atualizacoes, daemon=True).start()

    def report_callback_exception(self, exc, val, tb):
        # O .exe roda sem console (modo janela) — sem isso, qualquer exceção
        # dentro de um comando de botão, trace de variável ou bind seria
        # descartada silenciosamente (o comportamento padrão do Tkinter é
        # imprimir em stderr, que não existe nesse modo). Aqui ela vira log
        # visível tanto no arquivo de log quanto no painel de Execução.
        mensagem = "".join(traceback.format_exception(exc, val, tb))
        try:
            self.logger.error(f"Exceção não tratada em callback da interface:\n{mensagem}")
        except Exception:
            pass
        try:
            self._registrar_log("Erro interno ao processar a última ação — veja o log para detalhes.", nivel="error")
        except Exception:
            pass

    def _ao_fechar_janela(self):
        try:
            self.after_cancel(self._id_after_fila)
        except (ValueError, tk.TclError):
            pass
        if self._thread_geracao is not None and self._thread_geracao.is_alive():
            self._thread_geracao.join(timeout=5)
        self.logger.info("Aplicação encerrada pelo usuário.")
        self.quit()
        self.destroy()

    # ------------------------------------------------------------------
    # Montagem geral
    # ------------------------------------------------------------------

    def _definir_icone_janela(self):
        try:
            if os.path.exists(CAMINHO_LOGO_ICO):
                self.iconbitmap(CAMINHO_LOGO_ICO)
                return
        except tk.TclError:
            pass
        try:
            self._icone_janela = tk.PhotoImage(file=CAMINHO_LOGO_ICONE)
            self.iconphoto(True, self._icone_janela)
        except tk.TclError:
            pass

    def _montar_cabecalho(self, master):
        cabecalho = ttk.Frame(master)
        cabecalho.pack(fill="x", side="top")

        try:
            # Só a marca (sem texto) — o logo completo, reduzido a essa
            # altura, deixava o texto "2D CONSULTORES" ilegível/cortado.
            logo_imagem = tk.PhotoImage(file=CAMINHO_LOGO_ICONE)
            fator = max(1, logo_imagem.height() // 44)
            if fator > 1:
                logo_imagem = logo_imagem.subsample(fator, fator)
            self._logo_cabecalho_ref = logo_imagem
            ttk.Label(cabecalho, image=logo_imagem).pack(side="left", padx=(10, 10), pady=8)
        except tk.TclError:
            pass

        bloco_titulo = ttk.Frame(cabecalho)
        bloco_titulo.pack(side="left", pady=8)
        ttk.Label(bloco_titulo, text=NOME_SISTEMA, font=("Segoe UI", 13, "bold")).pack(anchor="w")
        ttk.Label(bloco_titulo, text=f"{NOME_EMPRESA} · v{VERSAO_ATUAL}", font=("Segoe UI", 9), foreground="gray").pack(anchor="w")

        self.label_boas_vindas = ttk.Label(cabecalho, text="", font=("Segoe UI", 10), foreground="#1565c0")
        self.label_boas_vindas.pack(side="right", padx=12)
        self._atualizar_boas_vindas()

        ttk.Separator(master, orient="horizontal").pack(fill="x", side="top")

    def _atualizar_boas_vindas(self):
        nome = (self.perfil or {}).get("nome", "").strip()
        self.label_boas_vindas.config(text=f"Bem-vindo, {nome}" if nome else "")

    def _montar_menu(self):
        barra_menu = tk.Menu(self)

        menu_arquivo = tk.Menu(barra_menu, tearoff=False)
        menu_arquivo.add_command(label="Selecionar CSV de vendas...", accelerator="Ctrl+O", command=self._selecionar_csv)
        menu_arquivo.add_command(label="Limpar base", command=self._limpar_base)
        menu_arquivo.add_separator()
        menu_arquivo.add_command(label="Buscar atualizações...", command=self._buscar_atualizacoes_manual)
        menu_arquivo.add_separator()
        menu_arquivo.add_command(label="Sair", accelerator="Ctrl+Q", command=self._ao_fechar_janela)
        barra_menu.add_cascade(label="Arquivo", menu=menu_arquivo)

        menu_relatorios = tk.Menu(barra_menu, tearoff=False)
        menu_relatorios.add_command(label="Gerar Relatório Padrão", accelerator="Ctrl+G", command=self._gerar_relatorio_padrao)
        menu_relatorios.add_command(label="Atualizar prévia dos grupos", accelerator="F5", command=self._pre_visualizar_grupos)
        barra_menu.add_cascade(label="Relatórios", menu=menu_relatorios)

        menu_ajuda = tk.Menu(barra_menu, tearoff=False)
        menu_ajuda.add_command(label="Sobre", accelerator="F1", command=self._mostrar_sobre)
        barra_menu.add_cascade(label="Ajuda", menu=menu_ajuda)

        self.config(menu=barra_menu)

        self.bind_all("<Control-o>", lambda evento: self._selecionar_csv())
        self.bind_all("<Control-q>", lambda evento: self._ao_fechar_janela())
        self.bind_all("<Control-g>", lambda evento: self._gerar_relatorio_padrao())
        self.bind_all("<F5>", lambda evento: self._pre_visualizar_grupos())
        self.bind_all("<F1>", lambda evento: self._mostrar_sobre())

    def _aplicar_estilos_customizados(self):
        estilo = ttk.Style(self)
        estilo.configure("Treeview", rowheight=26, font=("Segoe UI", 9))
        estilo.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        estilo.map("Treeview", background=[("selected", COR_ACCENT)], foreground=[("selected", "white")])

    def _atualizar_zebra_arvores(self):
        cor_par = "#eef2f6"
        cor_impar = "#ffffff"
        for arvore in (self.arvore_clientes, self.arvore_produtos):
            arvore.tag_configure("par", background=cor_par)
            arvore.tag_configure("impar", background=cor_impar)

    def _mostrar_sobre(self):
        messagebox.showinfo(
            "Sobre",
            f"{NOME_SISTEMA} — versão {VERSAO_ATUAL}\n{NOME_EMPRESA}\n\n"
            "Análise de funil de vendas B2B: tendências de produtos, erosão de\n"
            "clientes, segmentação por faturamento e relatórios personalizados.",
        )

    def _montar_interface(self):
        self._montar_menu()
        self._montar_cabecalho(self)

        corpo_principal = ttk.Frame(self)
        corpo_principal.pack(fill="both", expand=True)

        self.barra_lateral = ttk.Frame(corpo_principal, width=200)
        self.barra_lateral.pack(side="left", fill="y")
        self.barra_lateral.pack_propagate(False)

        area_conteudo = ttk.Frame(corpo_principal)
        area_conteudo.pack(side="left", fill="both", expand=True)
        area_conteudo.columnconfigure(0, weight=1)
        area_conteudo.rowconfigure(0, weight=1)

        self.paginas = {}
        self.botoes_nav = {}
        self.pagina_ativa = "Configurações"

        self.aba_configuracoes = ttk.Frame(area_conteudo)
        self._montar_aba_configuracoes(self.aba_configuracoes)
        self._registrar_pagina_nav("Configurações", "⚙", self.aba_configuracoes, area_conteudo)
        self._atualizar_zebra_arvores()

        self.aba_padrao = ttk.Frame(area_conteudo)
        self._montar_aba_padrao(self.aba_padrao)
        self._registrar_pagina_nav("Relatório Padrão", "📄", self.aba_padrao, area_conteudo)

        self.aba_personalizados = pb.ConstrutorRelatorioFrame(
            area_conteudo, obter_dataframe=lambda: self.df,
            relatorios_personalizados=self.relatorios_personalizados,
        )
        self._registrar_pagina_nav("Relatórios Personalizados", "📊", self.aba_personalizados, area_conteudo)

        self.aba_graficos = gf.PainelGraficosFrame(
            area_conteudo, obter_dataframe=lambda: self.df,
            obter_abc_df=self._obter_abc_df_atual,
        )
        self._registrar_pagina_nav("Gráficos", "📈", self.aba_graficos, area_conteudo)

        self.aba_perfil = ttk.Frame(area_conteudo)
        self._montar_aba_perfil(self.aba_perfil)
        self._registrar_pagina_nav("Perfil", "👤", self.aba_perfil, area_conteudo)

        self._montar_painel_execucao(self)

        self.status_var = tk.StringVar(value="Pronto.")
        barra_status = ttk.Label(self, textvariable=self.status_var, relief="sunken", anchor="w")
        barra_status.pack(fill="x", side="bottom")

        self._mostrar_pagina("Configurações")

    def _registrar_pagina_nav(self, nome, icone, frame_pagina, area_conteudo):
        frame_pagina.grid(in_=area_conteudo, row=0, column=0, sticky="nsew")
        self.paginas[nome] = frame_pagina

        rotulo = tk.Label(
            self.barra_lateral, text=f"   {icone}   {nome}", anchor="w",
            font=("Segoe UI", 10), padx=6, pady=12, cursor="hand2",
        )
        rotulo.pack(fill="x")
        rotulo.bind("<Button-1>", lambda evento: self._mostrar_pagina(nome))
        rotulo.bind("<Enter>", lambda evento: self._ao_passar_mouse_nav(nome, True))
        rotulo.bind("<Leave>", lambda evento: self._ao_passar_mouse_nav(nome, False))
        self.botoes_nav[nome] = rotulo

    def _mostrar_pagina(self, nome):
        self.pagina_ativa = nome
        self.paginas[nome].tkraise()
        self._atualizar_cores_nav()

    def _ao_passar_mouse_nav(self, nome, entrando):
        if nome == self.pagina_ativa:
            return
        rotulo = self.botoes_nav[nome]
        rotulo.configure(background=self._cor_hover_nav if entrando else self._cor_fundo_nav)

    def _atualizar_cores_nav(self):
        # Cores oficiais do tema sv_ttk claro (light.tcl) — o ttk.Style().lookup()
        # não retorna essas cores porque o tema é baseado em sprites de imagem,
        # não em "background" configurado por classe de estilo.
        self._cor_fundo_nav = "#fafafa"
        self._cor_texto_nav = "#1c1c1c"
        self._cor_hover_nav = "#d6e4f0"
        for nome, rotulo in self.botoes_nav.items():
            ativo = nome == self.pagina_ativa
            rotulo.configure(
                background=COR_ACCENT if ativo else self._cor_fundo_nav,
                foreground="white" if ativo else self._cor_texto_nav,
            )

    def _montar_painel_execucao(self, master):
        """
        Painel de log/progresso GLOBAL (visível em qualquer aba), não preso
        à aba Relatório Padrão — carregamento de CSV, geração de relatórios
        e qualquer outra tarefa de fundo aparecem aqui.
        """
        rodape = ttk.LabelFrame(master, text="Execução")
        rodape.pack(fill="x", side="bottom", padx=10, pady=(0, 4))

        self.progress = ttk.Progressbar(rodape, mode="indeterminate")
        self.progress.pack(fill="x", padx=8, pady=(8, 4))

        self.texto_log = tk.Text(rodape, height=6, state="disabled", wrap="word")
        self.texto_log.pack(fill="x", padx=8, pady=(0, 8))

    def _obter_abc_df_atual(self):
        if not self.resultados_analise:
            return None
        return self.resultados_analise.get(self.granularidade_referencia, {}).get("abc")

    def _definir_status(self, texto):
        self.status_var.set(texto)

    # ------------------------------------------------------------------
    # Zoom (aumenta/diminui a fonte padrão usada por toda a interface)
    # ------------------------------------------------------------------

    def _aplicar_zoom(self, delta=None):
        if delta is None:
            self.fator_zoom = 1.0
        else:
            self.fator_zoom = min(2.0, max(0.7, self.fator_zoom * delta))

        novo_tamanho = max(6, round(self.tamanho_fonte_base * self.fator_zoom))
        for nome_fonte in ("TkDefaultFont", "TkTextFont", "TkHeadingFont", "TkMenuFont", "TkFixedFont"):
            try:
                # root=self evita "Too early to use font: no default root window":
                # o TkinterDnD.Tk (usado para permitir drag-and-drop) nem sempre se
                # registra como root padrão do tkinter, então é preciso apontar
                # explicitamente qual interpretador Tcl usar.
                fonte = tkfont.nametofont(nome_fonte, root=self)
                fonte.configure(size=novo_tamanho)
            except (tk.TclError, RuntimeError):
                pass
        if hasattr(self, "label_zoom"):
            self.label_zoom.config(text=f"{int(self.fator_zoom * 100)}%")

    # ------------------------------------------------------------------
    # Aba "Configurações": base de dados, clientes/produtos, parâmetros
    # ------------------------------------------------------------------

    def _montar_aba_configuracoes(self, master):
        topo = ttk.Frame(master)
        topo.pack(fill="x", padx=10, pady=8)

        grupo_base = ttk.LabelFrame(topo, text="Base de dados")
        grupo_base.pack(side="left", fill="y")
        linha_base = ttk.Frame(grupo_base)
        linha_base.pack(padx=8, pady=6)
        ttk.Button(linha_base, text="Selecionar CSV de vendas...", command=self._selecionar_csv).pack(side="left")
        self.botao_limpar_base = ttk.Button(linha_base, text="Limpar base", command=self._limpar_base, state="disabled")
        self.botao_limpar_base.pack(side="left", padx=(6, 0))
        self.label_arquivo = ttk.Label(grupo_base, text="Nenhum arquivo carregado", foreground="gray")
        self.label_arquivo.pack(anchor="w", padx=8, pady=(0, 6))

        grupo_empresa = ttk.LabelFrame(topo, text="Empresa analisada")
        grupo_empresa.pack(side="left", fill="y", padx=8)
        self.entrada_nome_empresa = ttk.Entry(grupo_empresa, width=32)
        self.entrada_nome_empresa.pack(padx=8, pady=(6, 2))
        ttk.Label(
            grupo_empresa, text="Aparece na capa e no nome do arquivo", foreground="gray", font=("Segoe UI", 8),
        ).pack(padx=8, pady=(0, 6))

        grupo_config = ttk.LabelFrame(topo, text="Configuração da análise")
        grupo_config.pack(side="left", fill="y")
        linha_config = ttk.Frame(grupo_config)
        linha_config.pack(padx=8, pady=6)
        ttk.Button(linha_config, text="Salvar configuração...", command=self._salvar_configuracao_analise).pack(side="left")
        ttk.Button(linha_config, text="Carregar configuração...", command=self._carregar_configuracao_analise).pack(side="left", padx=(6, 0))

        grupo_zoom = ttk.LabelFrame(topo, text="Zoom")
        grupo_zoom.pack(side="right")
        linha_zoom = ttk.Frame(grupo_zoom)
        linha_zoom.pack(padx=8, pady=6)
        ttk.Button(linha_zoom, text="-", width=3, command=lambda: self._aplicar_zoom(1 / 1.1)).pack(side="left")
        self.label_zoom = ttk.Label(linha_zoom, text="100%", width=5, anchor="center")
        self.label_zoom.pack(side="left", padx=4)
        ttk.Button(linha_zoom, text="+", width=3, command=lambda: self._aplicar_zoom(1.1)).pack(side="left")
        ttk.Button(linha_zoom, text="Reset", command=lambda: self._aplicar_zoom(None)).pack(side="left", padx=(6, 0))

        corpo = ttk.Frame(master)
        corpo.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        coluna_parametros = self._montar_coluna_parametros(corpo)
        coluna_parametros.pack(side="left", fill="y", padx=(0, 8))

        coluna_clientes = self._montar_coluna_clientes(corpo)
        coluna_clientes.pack(side="left", fill="both", expand=True, padx=(0, 8))

        coluna_produtos = self._montar_coluna_produtos(corpo)
        coluna_produtos.pack(side="left", fill="both", expand=True)

    def _montar_coluna_parametros(self, master):
        frame = ttk.LabelFrame(master, text="Parâmetros (valem para todos os relatórios)")

        LARGURA_ROTULO = 26

        # --- Bloco 1: Segmentação (todos os inputs relacionados juntos,
        # rótulo e campo sempre nas mesmas duas colunas) -------------------
        bloco_segmentacao = ttk.LabelFrame(frame, text="Segmentação")
        bloco_segmentacao.pack(fill="x", padx=6, pady=(6, 6))

        linha = 0
        ttk.Button(bloco_segmentacao, text="Sugerir cortes automaticamente", command=self._sugerir_cortes_automaticamente).grid(
            row=linha, column=0, columnspan=2, sticky="we", padx=6, pady=(6, 2)
        )
        linha += 1
        self.botao_atualizar_preview = ttk.Button(bloco_segmentacao, text="Atualizar prévia dos grupos", command=self._pre_visualizar_grupos)
        self.botao_atualizar_preview.grid(row=linha, column=0, columnspan=2, sticky="we", padx=6, pady=2)
        linha += 1
        self.label_contagens_grupos = ttk.Label(
            bloco_segmentacao, text="Contagens: clique em 'Atualizar prévia' para calcular.",
            wraplength=230, foreground="#1565c0",
        )
        self.label_contagens_grupos.grid(row=linha, column=0, columnspan=2, sticky="w", padx=6, pady=(2, 6))
        linha += 1
        ttk.Separator(bloco_segmentacao, orient="horizontal").grid(
            row=linha, column=0, columnspan=2, sticky="we", padx=6, pady=(2, 6)
        )
        linha += 1

        ttk.Label(bloco_segmentacao, text="Corte de produtos por receita (%):", width=LARGURA_ROTULO, anchor="w").grid(
            row=linha, column=0, sticky="w", padx=6, pady=4
        )
        self.entrada_corte_produtos = ttk.Entry(bloco_segmentacao, width=8)
        self.entrada_corte_produtos.insert(0, "80")
        self.entrada_corte_produtos.bind("<KeyRelease>", lambda evento: self._marcar_configuracao_alterada())
        self.entrada_corte_produtos.grid(row=linha, column=1, sticky="w", padx=6, pady=4)
        linha += 1
        ttk.Label(bloco_segmentacao, text="padrão 80% — ajustável", foreground="gray", font=("Segoe UI", 8)).grid(
            row=linha, column=0, columnspan=2, sticky="w", padx=6
        )
        linha += 1

        self.entradas_corte_grupo = []
        for i, valor_padrao in enumerate(("30", "50", "60")):
            ttk.Label(bloco_segmentacao, text=f"Grupo {i + 1} de clientes até (%):", width=LARGURA_ROTULO, anchor="w").grid(
                row=linha, column=0, sticky="w", padx=6, pady=4
            )
            entrada = ttk.Entry(bloco_segmentacao, width=8)
            entrada.insert(0, valor_padrao)
            entrada.bind("<KeyRelease>", lambda evento: self._marcar_configuracao_alterada())
            entrada.grid(row=linha, column=1, sticky="w", padx=6, pady=4)
            self.entradas_corte_grupo.append(entrada)
            linha += 1

        ttk.Label(bloco_segmentacao, text="Máx. clientes por grupo:", width=LARGURA_ROTULO, anchor="w").grid(
            row=linha, column=0, sticky="w", padx=6, pady=4
        )
        self.entrada_max_por_grupo = ttk.Entry(bloco_segmentacao, width=8)
        self.entrada_max_por_grupo.insert(0, "10")
        self.entrada_max_por_grupo.grid(row=linha, column=1, sticky="w", padx=6, pady=4)
        linha += 1

        self.check_balcao = ttk.Checkbutton(
            bloco_segmentacao, text="Desconsiderar clientes balcão da frequência",
            command=self._marcar_configuracao_alterada,
        )
        self.check_balcao.state(["selected", "!alternate"])
        self.check_balcao.grid(row=linha, column=0, columnspan=2, sticky="w", padx=6, pady=(2, 0))
        linha += 1
        ttk.Label(
            bloco_segmentacao,
            text="Marcado, eles saem dos grupos e viram faixa \"Balcão\" — não\nsomem: filtre por \"Balcão\" na lista de clientes pra vê-los.",
            foreground="gray", font=("Segoe UI", 8), justify="left",
        ).grid(row=linha, column=0, columnspan=2, sticky="w", padx=6, pady=(0, 6))

        # --- Bloco 2: Evolução, alertas e erosão ---------------------------
        bloco_evolucao = ttk.LabelFrame(frame, text="Evolução, alertas e erosão")
        bloco_evolucao.pack(fill="x", padx=6, pady=(0, 6))

        self.check_incluir_periodo_atual = ttk.Checkbutton(
            bloco_evolucao, text="Incluir período mais recente (geralmente incompleto)",
            command=self._marcar_configuracao_alterada,
        )
        self.check_incluir_periodo_atual.state(["!selected", "!alternate"])
        self.check_incluir_periodo_atual.grid(row=0, column=0, columnspan=2, sticky="w", padx=6, pady=(6, 0))
        ttk.Label(
            bloco_evolucao, text="Desmarcado (padrão): o último período de cada relatório fica de fora.",
            foreground="gray", font=("Segoe UI", 8),
        ).grid(row=1, column=0, columnspan=2, sticky="w", padx=6, pady=(0, 6))

        ttk.Label(bloco_evolucao, text="Produtos a exibir (Evolução/Alertas):", width=LARGURA_ROTULO, anchor="w").grid(
            row=2, column=0, sticky="w", padx=6, pady=4
        )
        self.entrada_top_n_produtos = ttk.Entry(bloco_evolucao, width=8)
        self.entrada_top_n_produtos.grid(row=2, column=1, sticky="w", padx=6, pady=4)
        ttk.Label(bloco_evolucao, text="em branco = todos, ordenado por tendência", foreground="gray", font=("Segoe UI", 8)).grid(
            row=3, column=0, columnspan=2, sticky="w", padx=6
        )

        ttk.Label(bloco_evolucao, text="Redução mínima p/ erosão (%):", width=LARGURA_ROTULO, anchor="w").grid(
            row=4, column=0, sticky="w", padx=6, pady=(10, 4)
        )
        self.entrada_reducao_minima_erosao = ttk.Entry(bloco_evolucao, width=8)
        self.entrada_reducao_minima_erosao.insert(0, "50")
        self.entrada_reducao_minima_erosao.grid(row=4, column=1, sticky="w", padx=6, pady=(10, 4))
        ttk.Label(bloco_evolucao, text="só compara os 2 períodos mais recentes", foreground="gray", font=("Segoe UI", 8)).grid(
            row=5, column=0, columnspan=2, sticky="w", padx=6, pady=(0, 6)
        )

        # --- Bloco 3: Alertas e granularidade ------------------------------
        bloco_alertas = ttk.LabelFrame(frame, text="Alertas e granularidade")
        bloco_alertas.pack(fill="x", padx=6, pady=(0, 6))

        ttk.Label(bloco_alertas, text="Períodos seguidos em queda p/ alerta:", width=LARGURA_ROTULO, anchor="w").grid(
            row=0, column=0, sticky="w", padx=6, pady=4
        )
        self.entrada_periodos_queda = ttk.Entry(bloco_alertas, width=8)
        self.entrada_periodos_queda.insert(0, "2")
        self.entrada_periodos_queda.grid(row=0, column=1, sticky="w", padx=6, pady=4)
        ttk.Label(bloco_alertas, text="aplicado à granularidade escolhida abaixo", foreground="gray", font=("Segoe UI", 8)).grid(
            row=1, column=0, columnspan=2, sticky="w", padx=6
        )

        ttk.Label(bloco_alertas, text="Granularidade do relatório:", width=LARGURA_ROTULO, anchor="w").grid(
            row=2, column=0, sticky="w", padx=6, pady=(10, 4)
        )
        linha_granularidade = ttk.Frame(bloco_alertas)
        linha_granularidade.grid(row=3, column=0, columnspan=2, sticky="w", padx=6, pady=(0, 6))
        self.var_granularidade = tk.StringVar(value="Mensal")
        for granularidade in af.GRANULARIDADES:
            ttk.Radiobutton(
                linha_granularidade, text=granularidade, value=granularidade, variable=self.var_granularidade,
                command=lambda g=granularidade: self.var_granularidade.set(g),
            ).pack(side="left", padx=(0, 10))

        return frame

    def _montar_lista_com_filtro(self, master, titulo, colunas):
        """
        Cria um LabelFrame com campo de busca + filtro por grupo + Treeview
        (rola nativamente com a roda do mouse, ao contrário de uma lista de
        checkboxes).
        """
        frame = ttk.LabelFrame(master, text=titulo)

        barra_busca = ttk.Frame(frame)
        barra_busca.pack(fill="x", padx=6, pady=(6, 2))
        ttk.Label(barra_busca, text="Buscar:").pack(side="left")
        entrada_busca = ttk.Entry(barra_busca)
        entrada_busca.pack(side="left", fill="x", expand=True, padx=4)

        ttk.Label(barra_busca, text="Grupo:").pack(side="left", padx=(8, 2))
        combo_grupo = ttk.Combobox(barra_busca, state="readonly", width=10, values=["Todos"])
        combo_grupo.set("Todos")
        combo_grupo.pack(side="left")

        container = ttk.Frame(frame)
        container.pack(fill="both", expand=True, padx=6, pady=(2, 6))

        arvore = ttk.Treeview(container, columns=colunas, show="headings", selectmode="none")
        barra_rolagem = ttk.Scrollbar(container, orient="vertical", command=arvore.yview)
        arvore.configure(yscrollcommand=barra_rolagem.set)
        arvore.pack(side="left", fill="both", expand=True)
        barra_rolagem.pack(side="right", fill="y")

        return frame, entrada_busca, combo_grupo, arvore

    def _montar_coluna_clientes(self, master):
        frame, entrada_busca, combo_grupo, arvore = self._montar_lista_com_filtro(
            master, "Clientes a excluir das métricas", ("check", "cliente", "receita", "percentual", "grupo")
        )
        arvore.heading("check", text="Excluir?")
        arvore.heading("cliente", text="Cliente", command=lambda: self._ordenar_clientes("cliente"))
        arvore.heading("receita", text="Receita", command=lambda: self._ordenar_clientes("receita"))
        arvore.heading("percentual", text="% Receita", command=lambda: self._ordenar_clientes("percentual"))
        arvore.heading("grupo", text="Grupo")
        arvore.column("check", width=60, anchor="center")
        arvore.column("cliente", width=230, anchor="w")
        arvore.column("receita", width=100, anchor="e")
        arvore.column("percentual", width=75, anchor="e")
        arvore.column("grupo", width=75, anchor="center")

        arvore.bind("<Button-1>", lambda evento: self._alternar_checkbox_arvore(evento, arvore, self.estado_clientes))
        entrada_busca.bind("<KeyRelease>", lambda evento: self._buscar_clientes(entrada_busca.get()))
        combo_grupo.bind("<<ComboboxSelected>>", lambda evento: self._renderizar_clientes(entrada_busca.get()))

        self.arvore_clientes = arvore
        self.entrada_busca_clientes = entrada_busca
        self.combo_grupo_clientes = combo_grupo
        self.dados_clientes = []  # lista de dicts: cliente, receita, percentual, grupo
        return frame

    def _montar_coluna_produtos(self, master):
        frame, entrada_busca, combo_grupo, arvore = self._montar_lista_com_filtro(
            master, "Produtos considerados na análise", ("check", "produto", "grupo", "freq_simples", "freq_acumulado")
        )
        arvore.heading("check", text="Considerar?")
        arvore.heading("produto", text="Produto", command=self._ordenar_produtos)
        arvore.heading("grupo", text="Grupo")
        arvore.heading("freq_simples", text="Freq. Simples")
        arvore.heading("freq_acumulado", text="Freq. Acumulado")
        arvore.column("check", width=80, anchor="center")
        arvore.column("produto", width=230, anchor="w")
        arvore.column("grupo", width=70, anchor="center")
        arvore.column("freq_simples", width=90, anchor="e")
        arvore.column("freq_acumulado", width=100, anchor="e")

        arvore.bind("<Button-1>", lambda evento: self._alternar_checkbox_arvore(evento, arvore, self.estado_produtos))
        entrada_busca.bind("<KeyRelease>", lambda evento: self._buscar_produtos(entrada_busca.get()))
        combo_grupo.bind("<<ComboboxSelected>>", lambda evento: self._renderizar_produtos(entrada_busca.get()))

        self.arvore_produtos = arvore
        self.entrada_busca_produtos = entrada_busca
        self.combo_grupo_produtos = combo_grupo
        self.dados_produtos = []  # lista de dicts: chave, rotulo, grupo, freq_simples, freq_acumulado (% de receita)
        return frame

    def _alternar_checkbox_arvore(self, evento, arvore, dicionario_estado):
        if arvore.identify_region(evento.x, evento.y) != "cell":
            return
        coluna = arvore.identify_column(evento.x)
        linha = arvore.identify_row(evento.y)
        if not linha or coluna != "#1":
            return
        novo_valor = not dicionario_estado.get(linha, False)
        dicionario_estado[linha] = novo_valor
        valores = list(arvore.item(linha, "values"))
        valores[0] = "☑" if novo_valor else "☐"
        arvore.item(linha, values=valores)
        self._marcar_configuracao_alterada()

    def _marcar_configuracao_alterada(self):
        """Destaca visualmente o botão de prévia quando clientes/produtos mudam e a contagem fica desatualizada."""
        if hasattr(self, "botao_atualizar_preview"):
            self.botao_atualizar_preview.config(text="⚠ Atualizar prévia (parâmetros alterados)", style="Accent.TButton")

    # ------------------------------------------------------------------
    # Carregamento e limpeza do CSV
    # ------------------------------------------------------------------

    def _selecionar_csv(self):
        caminho = filedialog.askopenfilename(filetypes=[("CSV", "*.csv")])
        if not caminho:
            return

        self._definir_status("Carregando CSV...")
        self._registrar_log(f"Carregando CSV: {caminho}")
        self.progress.start(10)
        self.update_idletasks()
        try:
            self.df, linhas_vazias = af.carregar_csv(caminho)
            if linhas_vazias > 0:
                messagebox.showinfo("Linhas vazias ignoradas", f"{linhas_vazias} linha(s) com Ano ou Mês vazio ignorada(s).")
        except af.ErroCarregamentoCSV as exc:
            self.progress.stop()
            messagebox.showerror("Erro ao carregar CSV", str(exc))
            self._registrar_log(f"Falha ao carregar CSV: {exc}", nivel="error")
            self._definir_status("Falha ao carregar CSV.")
            return
        except Exception as exc:
            self.progress.stop()
            messagebox.showerror("Erro inesperado", f"Falha inesperada ao carregar o CSV:\n{exc}")
            self._registrar_log(f"Falha inesperada ao carregar CSV: {exc}", nivel="error")
            self._definir_status("Falha ao carregar CSV.")
            return
        self.progress.stop()

        self.label_arquivo.config(text=os.path.basename(caminho))
        self.botao_limpar_base.config(state="normal")
        self._popular_lista_clientes()
        self._popular_lista_produtos()
        self._registrar_log(f"CSV carregado com sucesso: {len(self.df)} linhas, "
                             f"{self.df['Cliente'].nunique()} clientes, {self.df['descricao'].nunique()} produtos.")
        self._definir_status(f"CSV carregado: {len(self.df)} linhas.")

    def _limpar_base(self):
        resposta = messagebox.askyesno("Limpar base", "Isso vai descartar o CSV carregado e todos os resultados. Continuar?")
        if not resposta:
            return

        self.df = None
        self.resultados_analise = None
        self.estado_clientes = {}
        self.estado_produtos = {}
        self.dados_clientes = []
        self.dados_produtos = []

        for linha in self.arvore_clientes.get_children():
            self.arvore_clientes.delete(linha)
        for linha in self.arvore_produtos.get_children():
            self.arvore_produtos.delete(linha)

        self.entrada_busca_clientes.delete(0, "end")
        self.entrada_busca_produtos.delete(0, "end")
        self.combo_grupo_clientes.set("Todos")
        self.combo_grupo_clientes["values"] = ["Todos"]
        self.combo_grupo_produtos.set("Todos")
        self.combo_grupo_produtos["values"] = ["Todos"]

        self.label_arquivo.config(text="Nenhum arquivo carregado")
        self.botao_limpar_base.config(state="disabled")
        self.label_contagens_grupos.config(text="Contagens: clique em 'Atualizar prévia' para calcular.")
        self.botao_atualizar_preview.config(text="Atualizar prévia dos grupos", style="TButton")

        self.texto_log.config(state="normal")
        self.texto_log.delete("1.0", "end")
        self.texto_log.config(state="disabled")

        self._registrar_log("Base de dados limpa pelo usuário. Sistema pronto para novo arquivo.")
        self._definir_status("Base limpa. Selecione um novo CSV.")

    def _popular_lista_clientes(self):
        receita_cliente = self.df.groupby("Cliente")["Receita"].sum().sort_values(ascending=False)
        total = receita_cliente.sum() or 1
        self.dados_clientes = [
            {"cliente": cliente, "receita": receita, "percentual": receita / total * 100, "grupo": "-"}
            for cliente, receita in receita_cliente.items()
        ]
        self.estado_clientes = {item["cliente"]: False for item in self.dados_clientes}
        self._ordem_clientes = ("receita", True)
        self._renderizar_clientes("")

    def _buscar_clientes(self, texto):
        # Buscar por nome deve achar o cliente esteja ele em qualquer grupo —
        # senão, procurar alguém que não está no grupo filtrado no momento
        # simplesmente nunca aparece (parece que "sumiu"/não foi encontrado).
        if texto.strip() and self.combo_grupo_clientes.get() != "Todos":
            self.combo_grupo_clientes.set("Todos")
        self._renderizar_clientes(texto)

    def _ordenar_clientes(self, campo):
        campo_atual, decrescente_atual = getattr(self, "_ordem_clientes", (None, True))
        decrescente = (not decrescente_atual) if campo == campo_atual else True
        self._ordem_clientes = (campo, decrescente)
        self.dados_clientes.sort(key=lambda item: item[campo], reverse=decrescente)
        self._renderizar_clientes(self.entrada_busca_clientes.get())

    def _renderizar_clientes(self, filtro):
        arvore = self.arvore_clientes
        for linha in arvore.get_children():
            arvore.delete(linha)
        filtro_lower = filtro.strip().lower()
        filtro_grupo = getattr(self, "combo_grupo_clientes", None)
        filtro_grupo = filtro_grupo.get() if filtro_grupo else "Todos"
        posicao = 0
        for item in self.dados_clientes:
            if filtro_lower and filtro_lower not in item["cliente"].lower():
                continue
            if filtro_grupo != "Todos" and item["grupo"] != filtro_grupo:
                continue
            marcado = self.estado_clientes.get(item["cliente"], False)
            arvore.insert("", "end", iid=item["cliente"], tags=("par" if posicao % 2 == 0 else "impar",), values=(
                "☑" if marcado else "☐",
                item["cliente"],
                _formatar_moeda_br(item["receita"]),
                f"{item['percentual']:.2f}%",
                item["grupo"],
            ))
            posicao += 1

    def _popular_lista_produtos(self):
        qtd_nao_harmonizados = af.contar_produtos_nao_harmonizados(self.df)
        produtos = sorted(p for p in self.df["descricao"].unique() if p != af.DESCRICAO_NAO_HARMONIZADA)

        self.dados_produtos = []
        if qtd_nao_harmonizados > 0:
            rotulo = f"⚠ {af.DESCRICAO_NAO_HARMONIZADA} ({qtd_nao_harmonizados} lançamentos sem descrição)"
            self.dados_produtos.append({"chave": af.DESCRICAO_NAO_HARMONIZADA, "rotulo": rotulo, "grupo": "-", "freq_simples": 0.0, "freq_acumulado": 0.0})
        self.dados_produtos.extend(
            {"chave": produto, "rotulo": produto, "grupo": "-", "freq_simples": 0.0, "freq_acumulado": 0.0}
            for produto in produtos
        )

        self.estado_produtos = {item["chave"]: True for item in self.dados_produtos}
        self._ordem_produtos = False
        self._recalcular_classificacoes()

    def _buscar_produtos(self, texto):
        # Mesma lógica de _buscar_clientes: busca por nome ignora o filtro de
        # grupo ativo, senão um produto fora do grupo filtrado nunca aparece.
        if texto.strip() and self.combo_grupo_produtos.get() != "Todos":
            self.combo_grupo_produtos.set("Todos")
        self._renderizar_produtos(texto)

    def _ordenar_produtos(self):
        decrescente = not getattr(self, "_ordem_produtos", False)
        self._ordem_produtos = decrescente
        self.dados_produtos.sort(key=lambda item: item["rotulo"].lower(), reverse=decrescente)
        self._renderizar_produtos(self.entrada_busca_produtos.get())

    def _renderizar_produtos(self, filtro):
        arvore = self.arvore_produtos
        for linha in arvore.get_children():
            arvore.delete(linha)
        filtro_lower = filtro.strip().lower()
        filtro_grupo = getattr(self, "combo_grupo_produtos", None)
        filtro_grupo = filtro_grupo.get() if filtro_grupo else "Todos"
        posicao = 0
        for item in self.dados_produtos:
            if filtro_lower and filtro_lower not in item["rotulo"].lower():
                continue
            if filtro_grupo != "Todos" and item["grupo"] != filtro_grupo:
                continue
            marcado = self.estado_produtos.get(item["chave"], True)
            arvore.insert("", "end", iid=item["chave"], tags=("par" if posicao % 2 == 0 else "impar",), values=(
                "☑" if marcado else "☐", item["rotulo"], item["grupo"],
                f"{item['freq_simples']:.2f}%", f"{item['freq_acumulado']:.2f}%",
            ))
            posicao += 1

    def _clientes_excluidos(self):
        return [cliente for cliente, excluido in self.estado_clientes.items() if excluido]

    def _produtos_excluidos(self):
        return [produto for produto, considerado in self.estado_produtos.items() if not considerado]

    # ------------------------------------------------------------------
    # Salvar/carregar configuração da análise (parâmetros + exclusões) em JSON
    # ------------------------------------------------------------------

    def _salvar_configuracao_analise(self):
        try:
            configuracao = {
                "corte_produtos": self.entrada_corte_produtos.get(),
                "cortes_grupo": [entrada.get() for entrada in self.entradas_corte_grupo],
                "max_por_grupo": self.entrada_max_por_grupo.get(),
                "periodos_queda": self.entrada_periodos_queda.get(),
                "granularidade": self.var_granularidade.get(),
                "clientes_excluidos": self._clientes_excluidos(),
                "produtos_excluidos": self._produtos_excluidos(),
                "formato_exportacao": getattr(self, "var_formato_exportacao", None) and self.var_formato_exportacao.get(),
            }
        except AttributeError:
            messagebox.showwarning("Salvar configuração", "A interface ainda não terminou de carregar.")
            return

        caminho = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("Configuração (JSON)", "*.json")])
        if not caminho:
            return
        with open(caminho, "w", encoding="utf-8") as arquivo:
            json.dump(configuracao, arquivo, ensure_ascii=False, indent=2)
        self._registrar_log(f"Configuração da análise salva em: {caminho}")
        messagebox.showinfo("Salvar configuração", "Configuração salva com sucesso.")

    def _carregar_configuracao_analise(self):
        caminho = filedialog.askopenfilename(filetypes=[("Configuração (JSON)", "*.json")])
        if not caminho:
            return
        try:
            with open(caminho, "r", encoding="utf-8") as arquivo:
                configuracao = json.load(arquivo)
        except Exception as exc:
            messagebox.showerror("Carregar configuração", f"Não foi possível ler o arquivo:\n{exc}")
            return

        self.entrada_corte_produtos.delete(0, "end")
        self.entrada_corte_produtos.insert(0, configuracao.get("corte_produtos", "80"))
        self._escrever_cortes_grupos(configuracao.get("cortes_grupo", ["30", "50", "60"]))
        self.entrada_max_por_grupo.delete(0, "end")
        self.entrada_max_por_grupo.insert(0, configuracao.get("max_por_grupo", "10"))
        self.entrada_periodos_queda.delete(0, "end")
        self.entrada_periodos_queda.insert(0, configuracao.get("periodos_queda", "2"))

        self.var_granularidade.set(configuracao.get("granularidade", "Mensal"))

        if hasattr(self, "var_formato_exportacao") and configuracao.get("formato_exportacao"):
            self.var_formato_exportacao.set(configuracao["formato_exportacao"])

        if self.df is not None:
            clientes_excluidos_salvos = set(configuracao.get("clientes_excluidos", []))
            for cliente in self.estado_clientes:
                self.estado_clientes[cliente] = cliente in clientes_excluidos_salvos

            produtos_excluidos_salvos = set(configuracao.get("produtos_excluidos", []))
            for produto in self.estado_produtos:
                self.estado_produtos[produto] = produto not in produtos_excluidos_salvos

            self._recalcular_classificacoes()
            messagebox.showinfo("Carregar configuração", "Configuração carregada e aplicada com sucesso.")
        else:
            messagebox.showinfo(
                "Carregar configuração",
                "Parâmetros carregados. Selecione o CSV para também aplicar clientes/produtos excluídos salvos.",
            )

        self._registrar_log(f"Configuração da análise carregada de: {caminho}")

    # ------------------------------------------------------------------
    # Parâmetros: sugestão e pré-visualização de grupos
    # ------------------------------------------------------------------

    def _ler_cortes_grupos(self):
        return [float(entrada.get().replace(",", ".")) for entrada in self.entradas_corte_grupo]

    def _escrever_cortes_grupos(self, cortes):
        for entrada, valor in zip(self.entradas_corte_grupo, cortes):
            entrada.delete(0, "end")
            entrada.insert(0, str(valor))

    def _sugerir_cortes_automaticamente(self):
        if self.df is None:
            messagebox.showwarning("Parâmetros", "Carregue um CSV antes de ajustar os parâmetros.")
            return
        try:
            cortes_iniciais = self._ler_cortes_grupos()
            max_por_grupo = int(self.entrada_max_por_grupo.get())
        except ValueError:
            messagebox.showerror("Parâmetros inválidos", "Verifique os campos numéricos dos grupos.")
            return

        cortes, contagens = af.sugerir_cortes_grupos(
            self.df, self._clientes_excluidos(), cortes_iniciais, max_por_grupo, desconsiderar_balcao=self.check_balcao.instate(["selected"])
        )
        self._escrever_cortes_grupos(cortes)
        self._registrar_log(f"Cortes sugeridos automaticamente: {cortes} (máx. {max_por_grupo} clientes/grupo).")
        self._recalcular_classificacoes()

        partes_resumo = "\n".join(
            f"  Grupo {i + 1}: até {corte}% da receita ({contagens[i]} clientes)"
            for i, corte in enumerate(cortes)
        )
        mudou = [round(a, 1) != round(b, 1) for a, b in zip(cortes_iniciais, cortes)]
        aviso_extra = ""
        if any(mudou):
            aviso_extra = (
                "\n\nOs percentuais foram reduzidos em relação ao que você tinha digitado —"
                " isso é esperado quando a receita está espalhada entre muitos clientes: para"
                " caber no máximo definido, o corte de % precisa ficar menor."
            )
        messagebox.showinfo(
            "Sugerir cortes automaticamente",
            f"Cortes ajustados para respeitar o máximo de {max_por_grupo} clientes por grupo:\n\n"
            f"{partes_resumo}\n  Demais: {contagens[-1]} clientes"
            f"{aviso_extra}",
        )

    def _pre_visualizar_grupos(self):
        self._recalcular_classificacoes()

    def _recalcular_classificacoes(self):
        """
        Recalcula, com os parâmetros ATUAIS da tela, a que grupo cada cliente
        e cada produto pertence (mais a frequência de compra dos produtos), e
        atualiza as colunas "Grupo"/"Frequência" nas duas listas — é a mesma
        classificação que rege os relatórios de segmentação, mas em versão
        rápida (agregada, não por período) só para a prévia na tela.
        """
        if self.df is None:
            return
        try:
            cortes = self._ler_cortes_grupos()
            corte_produtos = float(self.entrada_corte_produtos.get().replace(",", "."))
        except ValueError:
            messagebox.showerror("Parâmetros inválidos", "Verifique os campos numéricos de parâmetros.")
            return
        if any(cortes[i] >= cortes[i + 1] for i in range(len(cortes) - 1)) or any(c <= 0 or c >= 100 for c in cortes):
            messagebox.showerror("Parâmetros inválidos", "Os cortes de grupo devem ser crescentes e estar entre 0 e 100.")
            return
        if not (0 < corte_produtos < 100):
            messagebox.showerror("Parâmetros inválidos", "O corte de produtos deve estar entre 0 e 100.")
            return

        clientes_excluidos = set(self._clientes_excluidos())

        classificacao_clientes = af.classificar_clientes_agregado(self.df, clientes_excluidos, cortes, desconsiderar_balcao=self.check_balcao.instate(["selected"]))
        mapa_grupo_cliente = dict(zip(classificacao_clientes["Cliente"], classificacao_clientes["Faixa"]))
        mapa_percentual_cliente = dict(zip(classificacao_clientes["Cliente"], classificacao_clientes["Percentual_Individual"]))
        for item in self.dados_clientes:
            if item["cliente"] in clientes_excluidos:
                item["grupo"] = "Excluído"
                item["percentual"] = 0.0
            else:
                item["grupo"] = mapa_grupo_cliente.get(item["cliente"], "-")
                item["percentual"] = mapa_percentual_cliente.get(item["cliente"], 0.0)

        classificacao_produtos = af.classificar_produtos_agregado(self.df, corte_produtos)
        mapa_grupo_produto = dict(zip(classificacao_produtos["descricao"], classificacao_produtos["Faixa"]))
        mapa_freq_simples_produto = dict(zip(classificacao_produtos["descricao"], classificacao_produtos["Freq_Simples"]))
        mapa_freq_acumulado_produto = dict(zip(classificacao_produtos["descricao"], classificacao_produtos["Freq_Acumulado"]))
        for item in self.dados_produtos:
            item["grupo"] = mapa_grupo_produto.get(item["chave"], "-")
            item["freq_simples"] = mapa_freq_simples_produto.get(item["chave"], 0.0)
            item["freq_acumulado"] = mapa_freq_acumulado_produto.get(item["chave"], 0.0)

        nomes_grupos_clientes = [f"Grupo {i + 1}" for i in range(len(cortes))] + ["Demais", "Balcão", "Excluído"]
        if hasattr(self, "combo_grupo_clientes"):
            self.combo_grupo_clientes["values"] = ["Todos"] + nomes_grupos_clientes
            # Só troca o filtro sozinho se o usuário ainda não escolheu nada
            # específico ("Todos" = nenhuma preferência) — assim "Atualizar"
            # já mostra os clientes relevantes de cara, sem esconder uma
            # escolha manual que o usuário já tenha feito.
            if self.combo_grupo_clientes.get() == "Todos":
                self.combo_grupo_clientes.set("Grupo 1")
        if hasattr(self, "combo_grupo_produtos"):
            self.combo_grupo_produtos["values"] = ["Todos", "Grupo 1", "Demais"]
            if self.combo_grupo_produtos.get() == "Todos":
                self.combo_grupo_produtos.set("Grupo 1")

        contagens = af.contar_clientes_por_grupo(self.df, clientes_excluidos, cortes, desconsiderar_balcao=self.check_balcao.instate(["selected"]))
        self._exibir_contagens_grupos(cortes, contagens)

        self._renderizar_clientes(self.entrada_busca_clientes.get())
        self._renderizar_produtos(self.entrada_busca_produtos.get())

        self.botao_atualizar_preview.config(text="Atualizar prévia dos grupos", style="TButton")
        self._registrar_log(f"Prévia atualizada: cortes de grupo={cortes}, corte de produtos={corte_produtos}%.")

    def _exibir_contagens_grupos(self, cortes, contagens):
        partes = [f"Grupo {i + 1} (até {corte}%): {contagens[i]} clientes" for i, corte in enumerate(cortes)]
        partes.append(f"Demais: {contagens[-1]} clientes")
        self.label_contagens_grupos.config(text="  |  ".join(partes))

    # ------------------------------------------------------------------
    # Log de execução (arquivo + painel global na interface)
    # ------------------------------------------------------------------

    def _registrar_log(self, mensagem, nivel="info"):
        if nivel == "error":
            self.logger.error(mensagem)
        else:
            self.logger.info(mensagem)
        self.fila_eventos.put(("log", mensagem))

    def _bombear_fila_eventos(self):
        if not self.winfo_exists():
            return
        try:
            while True:
                # Um handler (ex.: fim de instalação de atualização) pode
                # fechar a janela no meio da leitura da fila — sem essa
                # checagem, o próximo item ainda seria processado contra
                # widgets já destruídos (TclError).
                if not self.winfo_exists():
                    return
                tipo, dados = self.fila_eventos.get_nowait()
                if tipo == "log":
                    self._anexar_log_ui(dados)
                elif tipo == "concluido":
                    self._ao_concluir_geracao(dados)
                elif tipo == "erro":
                    self._ao_falhar_geracao(dados)
                elif tipo == "atualizacao_checada":
                    self._ao_checar_atualizacao(dados)
                elif tipo == "atualizacao_progresso":
                    self._definir_status(f"Baixando atualização... {dados}%")
                elif tipo == "atualizacao_instalacao_erro":
                    self._ao_falhar_instalacao_atualizacao(dados)
                elif tipo == "atualizacao_instalacao_ok":
                    self._ao_concluir_instalacao_atualizacao()
        except queue.Empty:
            pass
        self._id_after_fila = self.after(150, self._bombear_fila_eventos)

    def _verificar_atualizacoes(self, manual=False):
        # Roda em thread separada (chamada de rede) — nunca deve travar a
        # abertura do sistema. A checagem automática do startup só interrompe
        # o usuário se houver versão nova (falha vira só uma linha no log,
        # pra dar pra diagnosticar depois); a manual (menu Arquivo > Buscar
        # atualizações) sempre mostra o resultado, incluindo erro.
        try:
            resultado = atualizacoes.verificar_nova_versao()
        except Exception as exc:
            self.fila_eventos.put(("atualizacao_checada", ("erro", str(exc), manual)))
            return
        if resultado:
            self.fila_eventos.put(("atualizacao_checada", ("disponivel", resultado, manual)))
        else:
            self.fila_eventos.put(("atualizacao_checada", ("atualizado", None, manual)))

    def _buscar_atualizacoes_manual(self):
        self._registrar_log("Verificando atualizações...")
        threading.Thread(target=self._verificar_atualizacoes, args=(True,), daemon=True).start()

    def _ao_checar_atualizacao(self, dados):
        status, info, manual = dados
        if status == "disponivel":
            tag, url_pagina, url_download_exe = info
            self._registrar_log(f"Nova versão disponível no GitHub: {tag} (versão atual: {VERSAO_ATUAL}).")

            if getattr(sys, "frozen", False) and url_download_exe:
                instalar = messagebox.askyesno(
                    "Nova versão disponível",
                    f"Há uma nova versão do {NOME_SISTEMA} disponível ({tag}).\n"
                    f"Versão instalada: {VERSAO_ATUAL}.\n\n"
                    "Deseja baixar e instalar agora? O programa vai fechar e abrir a versão "
                    "nova automaticamente.",
                )
                if instalar:
                    self._baixar_e_instalar_atualizacao(url_download_exe)
            else:
                abrir = messagebox.askyesno(
                    "Nova versão disponível",
                    f"Há uma nova versão do {NOME_SISTEMA} disponível ({tag}).\n"
                    f"Versão instalada: {VERSAO_ATUAL}.\n\n"
                    "Deseja abrir a página de download agora?",
                )
                if abrir:
                    webbrowser.open(url_pagina)
        elif status == "atualizado":
            self._registrar_log(f"Checagem de atualização: versão {VERSAO_ATUAL} já é a mais recente.")
            if manual:
                messagebox.showinfo("Buscar atualizações", f"Você já está na versão mais recente ({VERSAO_ATUAL}).")
        elif status == "erro":
            self._registrar_log(f"Falha ao checar atualizações: {info}", nivel="error")
            if manual:
                messagebox.showerror(
                    "Buscar atualizações",
                    f"Não foi possível verificar atualizações:\n{info}\n\n"
                    "Verifique a conexão com a internet (ou se há proxy/firewall bloqueando "
                    "acesso a api.github.com) e tente novamente.",
                )

    def _baixar_e_instalar_atualizacao(self, url_download_exe):
        self._registrar_log("Baixando atualização...")
        self._definir_status("Baixando atualização... 0%")
        threading.Thread(target=self._baixar_e_instalar_thread, args=(url_download_exe,), daemon=True).start()

    def _baixar_e_instalar_thread(self, url_download_exe):
        try:
            caminho_novo_exe = atualizacoes.baixar_atualizacao(
                url_download_exe,
                callback_progresso=lambda pct: self.fila_eventos.put(("atualizacao_progresso", pct)),
            )
            atualizacoes.aplicar_atualizacao_e_reiniciar(caminho_novo_exe)
        except Exception as exc:
            self.fila_eventos.put(("atualizacao_instalacao_erro", str(exc)))
            return
        self.fila_eventos.put(("atualizacao_instalacao_ok", None))

    def _ao_falhar_instalacao_atualizacao(self, erro):
        self._registrar_log(f"Falha ao instalar atualização automaticamente: {erro}", nivel="error")
        self._definir_status("Falha ao instalar atualização.")
        messagebox.showerror(
            "Atualização",
            f"Não foi possível instalar a atualização automaticamente:\n{erro}\n\n"
            "Baixe manualmente pela página de releases no GitHub.",
        )

    def _ao_concluir_instalacao_atualizacao(self):
        self._registrar_log("Atualização baixada. Reiniciando o programa...")
        self._definir_status("Reiniciando com a nova versão...")
        # Fecha em um ciclo separado do loop de eventos, não aqui direto:
        # destruir a janela no meio da leitura da fila (_bombear_fila_eventos)
        # faz o próximo item pendente (esse log, por exemplo) ser processado
        # contra widgets já destruídos.
        self.after(10, self._ao_fechar_janela)

    def _anexar_log_ui(self, mensagem):
        marca_horario = datetime.now().strftime("%H:%M:%S")
        self.texto_log.config(state="normal")
        self.texto_log.insert("end", f"[{marca_horario}] {mensagem}\n")
        self.texto_log.see("end")
        self.texto_log.config(state="disabled")

    # ------------------------------------------------------------------
    # Aba "Relatório Padrão": catálogo de relatórios prontos
    # ------------------------------------------------------------------

    def _montar_aba_padrao(self, master):
        ttk.Label(
            master, text="Selecione quais relatórios prontos deseja incluir no Excel final.",
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w", padx=10, pady=(10, 0))
        ttk.Label(
            master, text="Os parâmetros (base de dados, clientes/produtos, cortes) são definidos na aba Configurações.",
            foreground="gray",
        ).pack(anchor="w", padx=10, pady=(0, 8))

        barra_selecao = ttk.Frame(master)
        barra_selecao.pack(fill="x", padx=10)
        ttk.Button(barra_selecao, text="Selecionar todos", command=lambda: self._marcar_catalogo(True)).pack(side="left")
        ttk.Button(barra_selecao, text="Limpar seleção", command=lambda: self._marcar_catalogo(False)).pack(side="left", padx=6)
        self.label_contagem_catalogo = ttk.Label(barra_selecao, foreground="#1565c0")
        self.label_contagem_catalogo.pack(side="right")

        lista_frame = ttk.Frame(master)
        lista_frame.pack(fill="both", expand=True, padx=10, pady=8)
        lista_frame.columnconfigure(0, weight=1, uniform="categoria")
        lista_frame.columnconfigure(1, weight=1, uniform="categoria")

        # O estado de cada relatório é o próprio estado ttk do Checkbutton
        # ("selected"/"!selected"), sem tk.BooleanVar/trace — usar variável
        # ligada se mostrou pouco confiável (o clique às vezes não disparava
        # a escrita na variável). command= nunca falhou nos testes.
        self.checkboxes_catalogo = {}
        for i, (categoria, itens) in enumerate(CATALOGO_RELATORIOS):
            grupo = ttk.LabelFrame(lista_frame, text=categoria)
            grupo.grid(row=i // 2, column=i % 2, sticky="nsew", padx=6, pady=6)
            for j, (chave, titulo) in enumerate(itens):
                caixa = ttk.Checkbutton(grupo, text=titulo, command=self._atualizar_contagem_catalogo)
                caixa.state(["selected", "!alternate"])
                caixa.grid(row=j, column=0, sticky="w", padx=10, pady=4)
                self.checkboxes_catalogo[chave] = caixa
        self._atualizar_contagem_catalogo()

        area_exportacao = ttk.LabelFrame(master, text="Exportação")
        area_exportacao.pack(fill="x", padx=10, pady=(0, 10))

        barra_formato = ttk.Frame(area_exportacao)
        barra_formato.pack(pady=(10, 0))
        ttk.Label(barra_formato, text="Formato de exportação:").pack(side="left", padx=(0, 8))
        self.var_formato_exportacao = tk.StringVar(value="Excel")
        for formato in ("Excel", "PDF", "Word"):
            ttk.Radiobutton(
                barra_formato, text=formato, value=formato, variable=self.var_formato_exportacao,
                # command explícito além da variável ligada — o clique no radio nem sempre
                # escreve na variável de forma confiável (mesmo problema visto nas checkboxes
                # do catálogo), então forçamos aqui como garantia.
                command=lambda f=formato: self.var_formato_exportacao.set(f),
            ).pack(side="left", padx=4)
        ttk.Label(
            area_exportacao, text="PDF/Word mostram no máximo 50 linhas por tabela (formato de leitura, não de dados).\n"
                                  "Para a base completa (ex.: todos os clientes da segmentação ABC), use Excel.",
            justify="center", foreground="gray", font=("Segoe UI", 8),
        ).pack(pady=(4, 0))

        self.botao_gerar = ttk.Button(area_exportacao, text="Gerar Relatório Padrão", command=self._gerar_relatorio_padrao)
        self.botao_gerar.pack(pady=14)

    def _marcar_catalogo(self, valor):
        for caixa in self.checkboxes_catalogo.values():
            caixa.state(["selected", "!alternate"] if valor else ["!selected", "!alternate"])
        self._atualizar_contagem_catalogo()

    def _chaves_catalogo_selecionadas(self):
        return [chave for chave, caixa in self.checkboxes_catalogo.items() if caixa.instate(["selected"])]

    def _atualizar_contagem_catalogo(self):
        total = len(self.checkboxes_catalogo)
        selecionados = len(self._chaves_catalogo_selecionadas())
        self.label_contagem_catalogo.config(text=f"{selecionados} de {total} relatórios selecionados")

    # ------------------------------------------------------------------
    # Geração do relatório padrão (roda em thread para não travar a UI)
    # ------------------------------------------------------------------

    def _gerar_relatorio_padrao(self):
        if self.thread_em_execucao:
            messagebox.showinfo("Relatório", "Já existe um processamento em andamento.")
            return
        if self.df is None:
            messagebox.showwarning("Relatório", "Nenhuma base carregada. Vá em Configurações e selecione um CSV.")
            return

        chaves_selecionadas = self._chaves_catalogo_selecionadas()
        if not chaves_selecionadas:
            messagebox.showwarning("Relatório", "Selecione ao menos um relatório do catálogo.")
            return

        granularidades = [self.var_granularidade.get()]

        try:
            corte_produtos = float(self.entrada_corte_produtos.get().replace(",", "."))
            cortes_grupos = self._ler_cortes_grupos()
            periodos_queda = int(self.entrada_periodos_queda.get())
            texto_top_n = self.entrada_top_n_produtos.get().strip()
            top_n_produtos = int(texto_top_n) if texto_top_n else None
            reducao_minima_erosao = float(self.entrada_reducao_minima_erosao.get().replace(",", "."))
        except ValueError:
            messagebox.showerror("Parâmetros inválidos", "Verifique os campos numéricos em Configurações.")
            return

        if not (0 < corte_produtos < 100):
            messagebox.showerror("Parâmetros inválidos", "O corte de produtos deve estar entre 0 e 100.")
            return
        if any(cortes_grupos[i] >= cortes_grupos[i + 1] for i in range(len(cortes_grupos) - 1)) or \
           any(c <= 0 or c >= 100 for c in cortes_grupos):
            messagebox.showerror("Parâmetros inválidos", "Os cortes de grupo devem ser crescentes e estar entre 0 e 100.")
            return
        if top_n_produtos is not None and top_n_produtos <= 0:
            messagebox.showerror("Parâmetros inválidos", "Produtos a exibir deve ser um número positivo (ou em branco).")
            return
        if not (0 <= reducao_minima_erosao <= 100):
            messagebox.showerror("Parâmetros inválidos", "Redução mínima para erosão deve estar entre 0 e 100.")
            return

        produtos_excluidos = self._produtos_excluidos()
        df_filtrado = self.df[~self.df["descricao"].isin(produtos_excluidos)] if produtos_excluidos else self.df
        if df_filtrado.empty:
            messagebox.showerror("Relatório", "Nenhuma linha restante após excluir os produtos desmarcados.")
            return

        clientes_excluidos = self._clientes_excluidos()
        self.granularidade_referencia = granularidades[0]
        self._chaves_selecionadas_geracao = chaves_selecionadas
        self._formato_geracao = self.var_formato_exportacao.get()
        self._nome_empresa_geracao = self.entrada_nome_empresa.get().strip()

        self.thread_em_execucao = True
        self.botao_gerar.config(state="disabled")
        self.progress.start(10)
        self._registrar_log(
            f"Iniciando geração do relatório. Granularidades: {granularidades}. "
            f"Relatórios selecionados: {len(chaves_selecionadas)}. "
            f"Clientes excluídos: {len(clientes_excluidos)}. Produtos excluídos: {len(produtos_excluidos)}."
        )
        self._definir_status("Processando análises em segundo plano...")

        self._thread_geracao = threading.Thread(
            target=self._executar_geracao_em_thread,
            args=(df_filtrado, granularidades, clientes_excluidos, tuple(cortes_grupos), corte_produtos,
                  periodos_queda, set(chaves_selecionadas), self.check_balcao.instate(["selected"]),
                  not self.check_incluir_periodo_atual.instate(["selected"]), top_n_produtos, reducao_minima_erosao),
            daemon=True,
        )
        self._thread_geracao.start()

    def _executar_geracao_em_thread(self, df_filtrado, granularidades, clientes_excluidos,
                                     cortes_grupos, corte_produtos, periodos_queda, chaves_selecionadas, desconsiderar_balcao,
                                     excluir_periodo_atual, top_n_produtos, reducao_minima_erosao):
        try:
            resultados = af.gerar_analises_completas(
                df_filtrado, granularidades,
                clientes_excluidos=clientes_excluidos,
                cortes_clientes=cortes_grupos,
                corte_produtos=corte_produtos,
                periodos_queda_consecutiva=periodos_queda,
                chaves_solicitadas=chaves_selecionadas,
                callback_log=lambda mensagem: self._registrar_log(mensagem),
                desconsiderar_balcao=desconsiderar_balcao,
                excluir_periodo_atual=excluir_periodo_atual,
                top_n_produtos=top_n_produtos,
                reducao_minima_erosao=reducao_minima_erosao,
            )
            self.fila_eventos.put(("concluido", resultados))
        except Exception as exc:
            self.logger.exception("Falha ao processar análises")
            self.fila_eventos.put(("erro", str(exc)))

    def _ao_concluir_geracao(self, resultados):
        self.progress.stop()
        self.botao_gerar.config(state="normal")
        self.thread_em_execucao = False
        self.resultados_analise = resultados
        self._registrar_log("Análises concluídas com sucesso.")

        formato = self._formato_geracao
        opcoes_formato = {
            "Excel": (".xlsx", [("Excel", "*.xlsx")]),
            "PDF": (".pdf", [("PDF", "*.pdf")]),
            "Word": (".docx", [("Word", "*.docx")]),
        }
        extensao, tipos_arquivo = opcoes_formato[formato]

        nome_empresa = self._nome_empresa_geracao
        nome_empresa_arquivo = re.sub(r'[\\/*?:"<>|]', "", nome_empresa).strip()
        prefixo = f"Relatorio_{nome_empresa_arquivo}" if nome_empresa_arquivo else "Relatorio"
        nome_sugerido = f"{prefixo}_{datetime.now().strftime('%Y-%m')}"

        self._definir_status(f"Análises concluídas. Escolha onde salvar o {formato}.")
        caminho_saida = filedialog.asksaveasfilename(
            parent=self, defaultextension=extensao, filetypes=tipos_arquivo, initialfile=nome_sugerido
        )
        if not caminho_saida:
            self._registrar_log("Exportação cancelada pelo usuário.")
            self._definir_status("Relatório calculado, mas não exportado (cancelado pelo usuário).")
            return

        # gerar_analises_completas já só calcula/retorna as chaves pedidas
        # (ver chaves_solicitadas em _gerar_relatorio_padrao), então resultados
        # aqui já vem filtrado — nada a recortar de novo.
        resultados_filtrados = resultados

        self._definir_status(f"Exportando {formato}...")
        self._registrar_log(f"Exportando {formato} para: {caminho_saida}")
        try:
            if formato == "Excel":
                exportar_relatorio_excel(
                    caminho_saida, resultados_filtrados, self.relatorios_personalizados,
                    nome_usuario=self.perfil.get("nome", ""), nome_empresa=nome_empresa,
                )
            elif formato == "PDF":
                import exportadores_pdf_word
                exportadores_pdf_word.exportar_relatorio_pdf(
                    caminho_saida, resultados_filtrados, NOMES_ANALISE, nome_usuario=self.perfil.get("nome", ""),
                    colunas_moeda_por_analise=COLUNAS_MOEDA_POR_ANALISE, nome_empresa=nome_empresa,
                )
            else:
                import exportadores_pdf_word
                exportadores_pdf_word.exportar_relatorio_word(
                    caminho_saida, resultados_filtrados, NOMES_ANALISE, nome_usuario=self.perfil.get("nome", ""),
                    colunas_moeda_por_analise=COLUNAS_MOEDA_POR_ANALISE, nome_empresa=nome_empresa,
                )
        except Exception as exc:
            self.logger.exception(f"Falha ao exportar {formato}")
            messagebox.showerror(f"Erro ao exportar {formato}", str(exc))
            self._registrar_log(f"Falha ao exportar {formato}: {exc}", nivel="error")
            self._definir_status(f"Falha ao exportar {formato}.")
            return

        self._registrar_log("Relatório exportado com sucesso.")
        self._definir_status("Relatório gerado com sucesso.")
        messagebox.showinfo("Relatório", f"Relatório exportado com sucesso em:\n{caminho_saida}")

    def _ao_falhar_geracao(self, mensagem_erro):
        self.progress.stop()
        self.botao_gerar.config(state="normal")
        self.thread_em_execucao = False
        self._registrar_log(f"Falha ao processar análises: {mensagem_erro}", nivel="error")
        self._definir_status("Falha ao processar análises.")
        messagebox.showerror("Erro ao processar análises", mensagem_erro)

    # ------------------------------------------------------------------
    # Aba "Perfil": nome e preferências salvas localmente (SQLite)
    # ------------------------------------------------------------------

    def _montar_aba_perfil(self, master):
        frame = ttk.Frame(master)
        frame.pack(anchor="nw", padx=20, pady=20)

        ttk.Label(frame, text="Perfil do usuário", font=("Segoe UI", 13, "bold")).grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 12)
        )

        ttk.Label(frame, text="Seu nome:").grid(row=1, column=0, sticky="w", pady=4)
        self.entrada_nome_perfil = ttk.Entry(frame, width=32)
        self.entrada_nome_perfil.insert(0, self.perfil.get("nome", ""))
        self.entrada_nome_perfil.grid(row=1, column=1, sticky="w", padx=8, pady=4)

        ttk.Label(frame, text="Tamanho da fonte:").grid(row=2, column=0, sticky="w", pady=4)
        self.spin_tamanho_fonte = ttk.Spinbox(frame, from_=7, to=20, width=6)
        self.spin_tamanho_fonte.set(self.perfil.get("tamanho_fonte", perfil.TAMANHO_FONTE_PADRAO))
        self.spin_tamanho_fonte.grid(row=2, column=1, sticky="w", padx=8, pady=4)

        ttk.Button(frame, text="Salvar perfil", command=self._salvar_perfil).grid(
            row=3, column=0, columnspan=2, sticky="w", pady=(14, 4)
        )

        ttk.Label(
            frame,
            text="Estas preferências ficam salvas localmente (dados_locais/perfil.db), ao lado\n"
                 "do programa nesta máquina — não são incluídas se o executável for compartilhado.",
            foreground="gray", justify="left",
        ).grid(row=4, column=0, columnspan=2, sticky="w", pady=(10, 0))

    def _salvar_perfil(self):
        nome = self.entrada_nome_perfil.get().strip()
        try:
            tamanho_fonte = int(self.spin_tamanho_fonte.get())
        except ValueError:
            messagebox.showerror("Perfil", "Tamanho de fonte inválido.")
            return

        perfil.salvar_perfil(nome, tamanho_fonte)
        self.perfil = {"nome": nome, "tamanho_fonte": tamanho_fonte}
        self.tamanho_fonte_base = tamanho_fonte
        self._aplicar_zoom(None)
        self._atualizar_boas_vindas()
        self._registrar_log(f"Perfil salvo: nome='{nome}', tamanho_fonte={tamanho_fonte}.")
        messagebox.showinfo("Perfil", "Perfil salvo com sucesso.")


# ---------------------------------------------------------------------------
# Exportação para Excel
# ---------------------------------------------------------------------------

def _ajustar_largura_colunas(planilha):
    for coluna in planilha.columns:
        maior_comprimento = 0
        letra_coluna = get_column_letter(coluna[0].column)
        for celula in coluna:
            valor = str(celula.value) if celula.value is not None else ""
            maior_comprimento = max(maior_comprimento, len(valor))
        planilha.column_dimensions[letra_coluna].width = min(maior_comprimento + 2, 45)


def _formatar_cabecalho(planilha):
    preenchimento = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")
    fonte = Font(color="FFFFFF", bold=True)
    for celula in planilha[1]:
        celula.fill = preenchimento
        celula.font = fonte
        celula.alignment = Alignment(horizontal="center")


def _inserir_logo(planilha, coluna_ancora, linha_ancora=1, altura_pixels=34):
    """Insere a marca da 2D Consultores em miniatura, sem sobrepor os dados."""
    if ImagemExcel is None or not os.path.exists(CAMINHO_LOGO_ICONE):
        return
    try:
        imagem = ImagemExcel(CAMINHO_LOGO_ICONE)
        proporcao = imagem.width / imagem.height
        imagem.height = altura_pixels
        imagem.width = altura_pixels * proporcao
        planilha.add_image(imagem, f"{get_column_letter(coluna_ancora)}{linha_ancora}")
    except Exception:
        pass  # ausência da logo não deve impedir a geração do relatório


def _escrever_dataframe(workbook, nome_aba, df, colunas_moeda=None):
    if nome_aba in workbook.sheetnames:
        planilha = workbook[nome_aba]
    else:
        planilha = workbook.create_sheet(nome_aba)

    colunas_moeda = colunas_moeda or []
    df_para_exportar = df.reset_index() if df.index.name or isinstance(df.index, pd.MultiIndex) else df

    planilha.append(list(map(str, df_para_exportar.columns)))
    for _, linha in df_para_exportar.iterrows():
        planilha.append(list(linha))

    for indice_coluna, nome_coluna in enumerate(df_para_exportar.columns, start=1):
        if nome_coluna in colunas_moeda:
            for linha in range(2, planilha.max_row + 1):
                planilha.cell(row=linha, column=indice_coluna).number_format = 'R$ #,##0.00'

    _formatar_cabecalho(planilha)
    _ajustar_largura_colunas(planilha)
    _inserir_logo(planilha, coluna_ancora=len(df_para_exportar.columns) + 2)
    return planilha


def _criar_capa(workbook, resultados_analise, nome_usuario="", nome_empresa=""):
    """Primeira aba do relatório: logo, identidade da empresa e sumário do que foi gerado."""
    capa = workbook.create_sheet("Capa", 0)
    capa.sheet_view.showGridLines = False
    capa.column_dimensions["A"].width = 4
    capa.column_dimensions["B"].width = 60

    if os.path.exists(CAMINHO_LOGO) and ImagemExcel is not None:
        try:
            imagem = ImagemExcel(CAMINHO_LOGO)
            proporcao = imagem.width / imagem.height
            imagem.height = 130
            imagem.width = 130 * proporcao
            capa.add_image(imagem, "B2")
        except Exception:
            pass

    capa["B10"] = NOME_SISTEMA
    capa["B10"].font = Font(size=20, bold=True, color=COR_CABECALHO)
    capa["B11"] = NOME_EMPRESA
    capa["B11"].font = Font(size=12, color="666666")

    linha_info = 13
    if nome_empresa:
        capa[f"B{linha_info}"] = f"Empresa analisada: {nome_empresa}"
        capa[f"B{linha_info}"].font = Font(size=13, bold=True, color=COR_CABECALHO)
        linha_info += 1
    if nome_usuario:
        capa[f"B{linha_info}"] = f"Gerado por: {nome_usuario}"
        capa[f"B{linha_info}"].font = Font(size=10, italic=True, color="666666")
        linha_info += 1
    capa[f"B{linha_info}"] = f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    capa[f"B{linha_info}"].font = Font(size=10, italic=True, color="666666")

    linha = linha_info + 3
    capa[f"B{linha}"] = "Granularidades incluídas neste relatório:"
    capa[f"B{linha}"].font = Font(bold=True)
    for granularidade in resultados_analise.keys():
        linha += 1
        capa[f"B{linha}"] = f"•  {granularidade}"
    return capa


NOMES_ANALISE = {
    "top_produtos": "Top_Produtos",
    "top_clientes": "Top_Clientes",
    "top_fabricantes": "Top_Fabricantes",
    "poder_compra_clientes": "Poder_Compra_Clientes",
    "evolucao_produtos": "Evolucao_Produtos",
    "alertas_queda": "Alertas_Queda",
    "erosao_clientes": "Erosao_Clientes",
    "abc": "ABC_Clientes",
    "abc_produtos": "ABC_Produtos",
    "migracao_abc": "Migracao_ABC",
    "produtos_em_alta": "Produtos_Em_Alta",
    "produtos_em_queda": "Produtos_Em_Queda",
    "clientes_queda_qtd": "Clientes_Queda_Qtd",
    "correlacao_produto_cliente": "Correlacao_Prod_Cliente",
    "impacto_financeiro_churn": "Impacto_Financeiro_Churn",
}

COLUNAS_MOEDA_POR_ANALISE = {
    "top_produtos": ["Receita"],
    "top_clientes": ["Receita"],
    "top_fabricantes": ["Receita"],
    "poder_compra_clientes": ["Receita", "Renuncia", "Renuncia_Acumulada"],
    "evolucao_produtos": ["Receita", "Receita_Periodo_Anterior"],
    "alertas_queda": ["Receita_Ultimo_Periodo", "Receita_Primeiro_Periodo"],
    "erosao_clientes": ["Receita", "Receita_Periodo_Anterior", "Reducao_Receita"],
    "abc": ["Receita", "Renuncia", "Renuncia_Acumulada"],
    "abc_produtos": ["Receita", "Renuncia", "Renuncia_Acumulada"],
    "migracao_abc": [],
    "produtos_em_alta": ["Receita_Periodo_Anterior", "Receita_Periodo_Atual", "Total_Ano_Atual"],
    "produtos_em_queda": ["Receita_Periodo_Anterior", "Receita_Periodo_Atual", "Total_Ano_Atual"],
    "clientes_queda_qtd": ["Perda_Receita"],
    "correlacao_produto_cliente": ["Reducao_Receita"],
    "impacto_financeiro_churn": ["Receita_Sob_Risco"],
}


def exportar_relatorio_excel(caminho_saida, resultados_analise, relatorios_personalizados=None, nome_usuario="", nome_empresa=""):
    """
    Gera o arquivo .xlsx com uma aba por (análise x granularidade), formatado
    com cabeçalhos destacados, moeda BRL, largura de coluna automática e a
    logo da empresa em cada aba (mais uma capa de apresentação).
    """
    workbook = Workbook()
    workbook.remove(workbook.active)  # remove a aba padrão vazia
    _criar_capa(workbook, resultados_analise, nome_usuario, nome_empresa)

    for granularidade, analises in resultados_analise.items():
        for chave_analise, df_analise in analises.items():
            nome_base = NOMES_ANALISE.get(chave_analise, chave_analise)
            nome_aba = f"{nome_base}_{granularidade}"[:31]  # limite do Excel
            if df_analise is None or df_analise.empty:
                planilha = workbook.create_sheet(nome_aba)
                planilha.append(["Sem dados para esta análise/granularidade."])
                continue
            _escrever_dataframe(workbook, nome_aba, df_analise, COLUNAS_MOEDA_POR_ANALISE.get(chave_analise))

    if relatorios_personalizados:
        for nome_relatorio, tabela in relatorios_personalizados.items():
            nome_aba = f"Custom_{nome_relatorio}"[:31]
            _escrever_dataframe(workbook, nome_aba, tabela)

    if len(workbook.sheetnames) <= 1:
        workbook.create_sheet("Sem_Dados")

    workbook.save(caminho_saida)


def _pedir_permissao_primeira_execucao():
    """
    Na primeira vez que o programa roda nesta pasta (nem 'dados_locais' nem
    'logs' existem ainda ao lado do executável), pergunta ao usuário se pode
    criar essas pastas — perfil, configurações salvas e logs de execução.
    Se recusar, tudo isso vai para uma pasta temporária do Windows nesta
    sessão em vez de ficar ao lado do executável/script.
    """
    if recursos.pasta_dados_locais_ja_existe():
        return
    raiz_temporaria = tk.Tk()
    raiz_temporaria.withdraw()
    permitido = messagebox.askyesno(
        f"Primeira execução — {NOME_SISTEMA}",
        f"Esta é a primeira vez que o {NOME_SISTEMA} roda nesta pasta.\n\n"
        "Para funcionar, ele precisa criar, ao lado do programa:\n\n"
        "  •  uma pasta \"dados_locais\" — perfil do usuário e configurações salvas\n"
        "  •  uma pasta \"logs\" — registro de execução, para diagnóstico\n\n"
        "Nenhum dado sai desta máquina.\n\n"
        "Permitir a criação dessas pastas aqui?",
        icon="question",
    )
    raiz_temporaria.destroy()
    recursos.definir_permissao_dados_locais(permitido)


if __name__ == "__main__":
    import splash

    _pedir_permissao_primeira_execucao()
    splash.exibir_splash_e_iniciar(
        etapas_preparacao=construir_etapas_preparacao(),
        funcao_construir_janela_principal=AplicacaoAnaliseFunil,
    )
