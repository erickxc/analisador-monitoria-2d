"""
Geração do relatório em Excel (.xlsx): capa, uma aba por (análise x
granularidade), formatação de cabeçalho/moeda/percentual/largura de coluna e
a logo da empresa em cada aba.

Import pesado de propósito (pandas + openpyxl no topo do módulo) — diferente
de app.py, que adia esses imports pra não atrasar a splash. Este módulo só é
importado sob demanda, dentro de _ao_concluir_geracao quando o usuário escolhe
"Excel" (mesmo padrão já usado por exportadores_pdf_word.py para PDF/Word),
então pandas/openpyxl só carregam de fato na primeira exportação, não na
abertura do programa.
"""

import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ImagemExcel
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from catalogo import COR_CABECALHO, DESCRICAO_ANALISE, NOMES_ANALISE, _colunas_moeda_efetivas
from recursos import CAMINHO_LOGO, CAMINHO_LOGO_ICONE, NOME_EMPRESA, NOME_SISTEMA


def _ajustar_largura_colunas(planilha, ignorar_linhas=None):
    ignorar_linhas = ignorar_linhas or set()
    for coluna in planilha.columns:
        maior_comprimento = 0
        letra_coluna = get_column_letter(coluna[0].column)
        for celula in coluna:
            if celula.row in ignorar_linhas:
                continue
            valor = str(celula.value) if celula.value is not None else ""
            maior_comprimento = max(maior_comprimento, len(valor))
        planilha.column_dimensions[letra_coluna].width = min(maior_comprimento + 2, 45)


def _formatar_cabecalho(planilha, linha=1):
    preenchimento = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")
    fonte = Font(color="FFFFFF", bold=True)
    for celula in planilha[linha]:
        celula.fill = preenchimento
        celula.font = fonte
        celula.alignment = Alignment(horizontal="center")


def _escrever_descricao(planilha, descricao, n_colunas):
    """Linha de descrição (metodologia em 1-2 linhas) acima do cabeçalho da tabela, mesclada por toda a largura."""
    planilha.append([descricao])
    planilha.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(n_colunas, 1))
    celula = planilha.cell(row=1, column=1)
    celula.font = Font(italic=True, color=COR_CABECALHO)
    celula.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    planilha.row_dimensions[1].height = 28


def _inserir_logo(planilha, coluna_ancora, linha_ancora=1, altura_pixels=34):
    """Insere a marca da 2D Consultores em miniatura, sem sobrepor os dados."""
    if ImagemExcel is None or not os.path.exists(CAMINHO_LOGO_ICONE):
        return
    try:
        imagem = ImagemExcel(CAMINHO_LOGO_ICONE)
        proporcao = imagem.width / imagem.height
        imagem.height = altura_pixels
        imagem.width = altura_pixels * proporcao
        planilha.add_image(imagem, f"{get_column_letter(coluna_ancora)}{linha_ancora}")
    except Exception:
        pass  # ausência da logo não deve impedir a geração do relatório


def _eh_coluna_percentual(nome_coluna):
    """
    Detecta coluna de percentual pelo nome (Percentual_*, *_Pct, "% ..."),
    sem precisar de uma lista mantida à parte por análise — cobre também as
    colunas dos Relatórios Personalizados, que o usuário monta livremente.
    """
    nome = str(nome_coluna).lower()
    return "%" in nome or "percentual" in nome or "pct" in nome


def _escrever_dataframe(workbook, nome_aba, df, colunas_moeda=None, descricao=None):
    if nome_aba in workbook.sheetnames:
        planilha = workbook[nome_aba]
    else:
        planilha = workbook.create_sheet(nome_aba)

    colunas_moeda = colunas_moeda or []
    df_para_exportar = df.reset_index() if df.index.name or isinstance(df.index, pd.MultiIndex) else df
    colunas_percentual = [c for c in df_para_exportar.columns if _eh_coluna_percentual(c)]

    if descricao:
        _escrever_descricao(planilha, descricao, len(df_para_exportar.columns))
    linha_cabecalho = planilha.max_row + 1

    planilha.append(list(map(str, df_para_exportar.columns)))
    for _, linha in df_para_exportar.iterrows():
        valores = list(linha)
        # O motor de análise guarda percentual como número "cru" (8.3 = 8,3%).
        # O formato nativo de % do Excel multiplica por 100 na exibição, então
        # o valor gravado precisa ser a fração (0.083) — sem isso, a célula
        # mostra "830%" em vez de "8,30%" quando formatada como percentual.
        for indice, nome_coluna in enumerate(df_para_exportar.columns):
            if nome_coluna in colunas_percentual and pd.notnull(valores[indice]):
                valores[indice] = valores[indice] / 100
        planilha.append(valores)

    for indice_coluna, nome_coluna in enumerate(df_para_exportar.columns, start=1):
        if nome_coluna in colunas_moeda:
            for linha in range(linha_cabecalho + 1, planilha.max_row + 1):
                planilha.cell(row=linha, column=indice_coluna).number_format = 'R$ #,##0.00'
        elif nome_coluna in colunas_percentual:
            for linha in range(linha_cabecalho + 1, planilha.max_row + 1):
                planilha.cell(row=linha, column=indice_coluna).number_format = '0.00%'

    _formatar_cabecalho(planilha, linha=linha_cabecalho)
    _ajustar_largura_colunas(planilha, ignorar_linhas={1} if descricao else None)
    _inserir_logo(planilha, coluna_ancora=len(df_para_exportar.columns) + 2)
    return planilha


def _criar_capa(workbook, resultados_analise, nome_usuario="", nome_empresa=""):
    """Primeira aba do relatório: logo, identidade da empresa e sumário do que foi gerado."""
    capa = workbook.create_sheet("Capa", 0)
    capa.sheet_view.showGridLines = False
    capa.column_dimensions["A"].width = 4
    capa.column_dimensions["B"].width = 60

    if os.path.exists(CAMINHO_LOGO) and ImagemExcel is not None:
        try:
            imagem = ImagemExcel(CAMINHO_LOGO)
            proporcao = imagem.width / imagem.height
            imagem.height = 130
            imagem.width = 130 * proporcao
            capa.add_image(imagem, "B2")
        except Exception:
            pass

    capa["B10"] = NOME_SISTEMA
    capa["B10"].font = Font(size=20, bold=True, color=COR_CABECALHO)
    capa["B11"] = NOME_EMPRESA
    capa["B11"].font = Font(size=12, color="666666")

    linha_info = 13
    if nome_empresa:
        capa[f"B{linha_info}"] = f"Empresa analisada: {nome_empresa}"
        capa[f"B{linha_info}"].font = Font(size=13, bold=True, color=COR_CABECALHO)
        linha_info += 1
    if nome_usuario:
        capa[f"B{linha_info}"] = f"Gerado por: {nome_usuario}"
        capa[f"B{linha_info}"].font = Font(size=10, italic=True, color="666666")
        linha_info += 1
    capa[f"B{linha_info}"] = f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    capa[f"B{linha_info}"].font = Font(size=10, italic=True, color="666666")

    linha = linha_info + 3
    capa[f"B{linha}"] = "Granularidades incluídas neste relatório:"
    capa[f"B{linha}"].font = Font(bold=True)
    for granularidade in resultados_analise.keys():
        linha += 1
        capa[f"B{linha}"] = f"•  {granularidade}"
    return capa


def exportar_relatorio_excel(caminho_saida, resultados_analise, relatorios_personalizados=None, nome_usuario="", nome_empresa="", descricao_analise=None):
    """
    Gera o arquivo .xlsx com uma aba por (análise x granularidade), formatado
    com cabeçalhos destacados, moeda BRL, largura de coluna automática e a
    logo da empresa em cada aba (mais uma capa de apresentação).

    descricao_analise: dict chave->texto pra sobrepor DESCRICAO_ANALISE (ver
    catalogo._construir_descricoes_dinamicas) — usa os valores padrão se None.
    """
    descricoes = descricao_analise if descricao_analise is not None else DESCRICAO_ANALISE
    workbook = Workbook()
    workbook.remove(workbook.active)  # remove a aba padrão vazia
    _criar_capa(workbook, resultados_analise, nome_usuario, nome_empresa)

    for granularidade, analises in resultados_analise.items():
        for chave_analise, df_analise in analises.items():
            nome_base = NOMES_ANALISE.get(chave_analise, chave_analise)
            nome_aba = f"{nome_base}_{granularidade}"[:31]  # limite do Excel
            if df_analise is None or df_analise.empty:
                planilha = workbook.create_sheet(nome_aba)
                planilha.append(["Sem dados para esta análise/granularidade."])
                continue
            _escrever_dataframe(
                workbook, nome_aba, df_analise, _colunas_moeda_efetivas(chave_analise, df_analise),
                descricao=descricoes.get(chave_analise),
            )

    if relatorios_personalizados:
        for nome_relatorio, tabela in relatorios_personalizados.items():
            nome_aba = f"Custom_{nome_relatorio}"[:31]
            _escrever_dataframe(workbook, nome_aba, tabela)

    if len(workbook.sheetnames) <= 1:
        workbook.create_sheet("Sem_Dados")

    workbook.save(caminho_saida)
